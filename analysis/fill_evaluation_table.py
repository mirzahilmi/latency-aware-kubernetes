"""Fill TABEL_PARAMETER_EVALUASI.xlsx from the 24 k6 CSV outputs.

Summary statistics chosen by the user (applied uniformly across all 24 files,
recorded here per INSTRUCTION.md):

    WAKTU RESPON                 (http_req_duration) -> arithmetic MEAN of metric_value           [ms]
    THROUGHPUT                   (http_reqs)         -> ROUND(COUNT / (max(timestamp) - min(timestamp) + 1)) [req/s, integer]
    TINGKAT KEGAGALAN REQUEST    (http_req_failed)   -> "<100 * MEAN of metric_value> / <COUNT(*)>" [percentage / total requests, string]

All three statistics match k6's own end-of-run summary conventions.

Pipeline:
  - DuckDB reads all 24 CSVs in a single streamed scan per metric (three queries total).
  - Filename is parsed in SQL (regex) into (configuration, testcase) so no Python-side
    pandas/pd.read_csv ever touches the ~30 GB of input.
  - Only the resulting 24-row x 3-metric frame reaches Python.
  - Template rows are matched at runtime by normalizing column A so the script
    survives template edits (no hardcoded row numbers).

Output: TABEL_PARAMETER_EVALUASI.xlsx alongside the template. Template is never modified.
"""

from __future__ import annotations

import re
import sys
from pathlib import Path

import duckdb
import openpyxl

WORKING_DIRECTORY = Path(__file__).resolve().parent
TEMPLATE_PATH = WORKING_DIRECTORY / "TEMPLATE_TABEL_PARAMETER_EVALUASI.xlsx"
OUTPUT_PATH = WORKING_DIRECTORY / "TABEL_PARAMETER_EVALUASI.xlsx"
CSV_GLOB = str(WORKING_DIRECTORY / "RPS_DATASET_*_TESTCASE_*.csv")

FILENAME_PATTERN = r"RPS_DATASET_(BASELINE|IPVS_LC|SOLUTION)_TESTCASE_(\d+)\.csv"

# Test-case number -> RPS ratio string. Documented in CLAUDE.md.
# Used only for matching against the ratio string read from column A of the template
# at runtime — the actual row index is discovered, never hardcoded.
TESTCASE_TO_RATIO = {
    1: "200:0:0:0",
    2: "1600:0:0:0",
    3: "200:200:200:200",
    4: "800:800:800:800",
    5: "800:400:400:400",
    6: "3200:400:400:400",
    7: "1600:800:800:400",
    8: "1200:800:800:800",
}

# Filename configuration token -> destination template column letter.
CONFIGURATION_TO_COLUMN = {
    "SOLUTION": "B",  # EWMA       -> proposed solution
    "BASELINE": "C",  # BASELINE
    "IPVS_LC":  "D",  # LEASTCON
}

# Sheet name -> (k6 metric_name, human label for stdout printout)
SHEET_TO_METRIC = {
    "WAKTU RESPON":              ("http_req_duration", "Mean response time (ms)"),
    "THROUGHPUT":                ("http_reqs",         "Throughput (req/s)"),
    "TINGKAT KEGAGALAN REQUEST": ("http_req_failed",   "Failure rate (%)"),
}

# Force the dot decimal separator regardless of the viewer's regional settings.
# The [$-en-US] locale tag tells Excel/LibreOffice to render this cell using
# US English number conventions (dot decimal, comma thousands), overriding the
# default system locale (e.g. Indonesian, which would otherwise display 11,63 instead of 11.63).
NUMBER_FORMAT_DOT_DECIMAL = "[$-en-US]0.000000"
# Throughput is reported as whole requests per second — request counts divided by run duration
# happen to be fractional in math but conceptually represent a per-second event rate.
NUMBER_FORMAT_INTEGER = "[$-en-US]0"

NUMBER_FORMAT_TEXT = "@"

SHEET_TO_NUMBER_FORMAT = {
    "WAKTU RESPON":              NUMBER_FORMAT_DOT_DECIMAL,
    "THROUGHPUT":                NUMBER_FORMAT_INTEGER,
    "TINGKAT KEGAGALAN REQUEST": NUMBER_FORMAT_TEXT,
}


def log(message: str) -> None:
    print(message, file=sys.stderr, flush=True)


def normalize_ratio(cell_value: object) -> str | None:
    """Extract the four-component RPS ratio from any text representation.

    Returns the digits joined with ':', or None if the cell does not contain
    exactly four integer fields (which filters out headers and unit rows).
    """
    if cell_value is None:
        return None
    digits = re.findall(r"\d+", str(cell_value))
    if len(digits) != 4:
        return None
    return ":".join(digits)


def discover_template_rows(workbook: openpyxl.Workbook) -> dict[str, dict[str, int]]:
    """For each evaluation sheet, scan column A and return {normalized_ratio: row_index}."""
    sheet_row_maps: dict[str, dict[str, int]] = {}
    for sheet_name in SHEET_TO_METRIC:
        worksheet = workbook[sheet_name]
        ratio_to_row: dict[str, int] = {}
        for row_index in range(1, worksheet.max_row + 1):
            ratio = normalize_ratio(worksheet.cell(row=row_index, column=1).value)
            if ratio is not None:
                ratio_to_row[ratio] = row_index
        sheet_row_maps[sheet_name] = ratio_to_row
        log(f"  '{sheet_name}': matched {len(ratio_to_row)} ratio rows -> {ratio_to_row}")
    return sheet_row_maps


def aggregate_response_time_mean(connection: duckdb.DuckDBPyConnection) -> list[tuple]:
    log("Aggregating http_req_duration (mean of metric_value, ms)...")
    return connection.execute(
        rf"""
        SELECT
            regexp_extract(filename, '{FILENAME_PATTERN}', 1)                  AS configuration,
            CAST(regexp_extract(filename, '{FILENAME_PATTERN}', 2) AS INTEGER) AS testcase,
            AVG(metric_value)                                                  AS value
        FROM read_csv_auto('{CSV_GLOB}', filename=true, header=true)
        WHERE metric_name = 'http_req_duration'
        GROUP BY configuration, testcase
        ORDER BY configuration, testcase
        """
    ).fetchall()


def aggregate_throughput(connection: duckdb.DuckDBPyConnection) -> list[tuple]:
    log("Aggregating http_reqs (count / span, req/s)...")
    return connection.execute(
        rf"""
        SELECT
            regexp_extract(filename, '{FILENAME_PATTERN}', 1)                  AS configuration,
            CAST(regexp_extract(filename, '{FILENAME_PATTERN}', 2) AS INTEGER) AS testcase,
            CAST(
            ROUND(
                CAST(COUNT(*) AS DOUBLE)
                / (CAST(MAX(timestamp) AS BIGINT) - CAST(MIN(timestamp) AS BIGINT) + 1)
            ) AS BIGINT
        )                                                                      AS value
        FROM read_csv_auto('{CSV_GLOB}', filename=true, header=true)
        WHERE metric_name = 'http_reqs'
        GROUP BY configuration, testcase
        ORDER BY configuration, testcase
        """
    ).fetchall()


def aggregate_failure_rate(connection: duckdb.DuckDBPyConnection) -> list[tuple]:
    log("Aggregating http_req_failed (\"<percentage> / <total requests>\")...")
    rows = connection.execute(
        rf"""
        SELECT
            regexp_extract(filename, '{FILENAME_PATTERN}', 1)                  AS configuration,
            CAST(regexp_extract(filename, '{FILENAME_PATTERN}', 2) AS INTEGER) AS testcase,
            100.0 * AVG(metric_value)                                          AS percentage,
            COUNT(*)                                                           AS total_requests
        FROM read_csv_auto('{CSV_GLOB}', filename=true, header=true)
        WHERE metric_name = 'http_req_failed'
        GROUP BY configuration, testcase
        ORDER BY configuration, testcase
        """
    ).fetchall()
    def format_cell(percentage: float, total_requests: int) -> str:
        truncated_percentage = int(percentage * 1000) / 1000
        percentage_text = f"{truncated_percentage:.3f}"
        total_text = f"{total_requests:,}".replace(",", ".")
        return f"{percentage_text} / {total_text}"

    return [
        (configuration, testcase, format_cell(percentage, int(total_requests)))
        for configuration, testcase, percentage, total_requests in rows
    ]


SHEET_TO_AGGREGATOR = {
    "WAKTU RESPON":              aggregate_response_time_mean,
    "THROUGHPUT":                aggregate_throughput,
    "TINGKAT KEGAGALAN REQUEST": aggregate_failure_rate,
}


def write_values_into_workbook(
    workbook: openpyxl.Workbook,
    sheet_row_maps: dict[str, dict[str, int]],
    results_by_sheet: dict[str, dict[tuple[str, int], float]],
) -> None:
    for sheet_name in SHEET_TO_METRIC:
        worksheet = workbook[sheet_name]
        ratio_to_row = sheet_row_maps[sheet_name]
        results = results_by_sheet[sheet_name]
        for testcase_number, ratio in TESTCASE_TO_RATIO.items():
            normalized = normalize_ratio(ratio)
            if normalized not in ratio_to_row:
                log(f"  WARNING: ratio '{ratio}' not found in column A of '{sheet_name}'")
                continue
            target_row = ratio_to_row[normalized]
            for configuration, column_letter in CONFIGURATION_TO_COLUMN.items():
                value = results.get((configuration, testcase_number))
                if value is None:
                    log(f"  WARNING: missing aggregate for {configuration} TC{testcase_number} in '{sheet_name}'")
                    continue
                cell = worksheet[f"{column_letter}{target_row}"]
                number_format = SHEET_TO_NUMBER_FORMAT[sheet_name]
                if number_format == NUMBER_FORMAT_INTEGER:
                    cell.value = int(round(value))
                elif number_format == NUMBER_FORMAT_TEXT:
                    cell.value = str(value)
                else:
                    cell.value = float(value)
                cell.number_format = number_format


def print_spotcheck_tables(results_by_sheet: dict[str, dict[tuple[str, int], float]]) -> None:
    for sheet_name, (metric_name, label) in SHEET_TO_METRIC.items():
        print()
        print(f"=== {sheet_name} -- {metric_name} -- {label} ===")
        header = (
            f"{'TC':<4} {'RPS ratio (A:B:C:D)':<22} "
            f"{'EWMA (col B)':>16} {'BASELINE (col C)':>18} {'LEASTCON (col D)':>18}"
        )
        print(header)
        print("-" * len(header))
        results = results_by_sheet[sheet_name]
        for testcase_number in sorted(TESTCASE_TO_RATIO):
            ratio = TESTCASE_TO_RATIO[testcase_number]
            ewma     = results.get(("SOLUTION", testcase_number))
            baseline = results.get(("BASELINE", testcase_number))
            leastcon = results.get(("IPVS_LC",  testcase_number))
            def render(value: object) -> str:
                if value is None:
                    return "n/a"
                if isinstance(value, (int, float)):
                    return f"{value:.6f}"
                return str(value)
            print(
                f"TC{testcase_number:<2} {ratio:<22} "
                f"{render(ewma):>16} {render(baseline):>18} {render(leastcon):>18}"
            )


def main() -> int:
    if not TEMPLATE_PATH.exists():
        log(f"ERROR: template not found at {TEMPLATE_PATH}")
        return 1

    log(f"Loading template: {TEMPLATE_PATH.name}")
    workbook = openpyxl.load_workbook(TEMPLATE_PATH)
    log("Discovering ratio->row mapping from each sheet's column A...")
    sheet_row_maps = discover_template_rows(workbook)

    log("Opening DuckDB in-memory connection...")
    connection = duckdb.connect()

    results_by_sheet: dict[str, dict[tuple[str, int], float]] = {}
    for sheet_name in SHEET_TO_METRIC:
        rows = SHEET_TO_AGGREGATOR[sheet_name](connection)
        results_by_sheet[sheet_name] = {
            (configuration, testcase): value
            for configuration, testcase, value in rows
        }
        log(f"  '{sheet_name}': received {len(rows)} (configuration, testcase) pairs")

    connection.close()

    log("Writing values into the in-memory workbook copy...")
    write_values_into_workbook(workbook, sheet_row_maps, results_by_sheet)

    log(f"Saving filled workbook to: {OUTPUT_PATH.name}")
    workbook.save(OUTPUT_PATH)

    print_spotcheck_tables(results_by_sheet)
    log("Done.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
